import openpyxl
hdanda = openpyxl.load_workbook('HDA_NDA_Format.xlsx')
f=open('HDA.csv','r'); f1=open('NDA.csv','r')
hdadata=f.readlines(); ndadata=f1.readlines()
hdasheet = hdanda['HDA']; ndasheet=hdanda['NDA']
totalkm=0; runduty=0; fixedsignon=0; A=0; B=0; C=0; D=0
month=input('Enter month name: ')
ndasheet['I1']=hdasheet['G1']='MONTH: '+month
for i in range(len(hdadata)-1):
	for j in range(6):
		hdasheet.cell(row=13+i,column=j+1).value=hdadata[i+1].split(',')[j]
	totalkm+=float(hdadata[i+1].split(',')[5])
	if hdadata[i+1].split(',')[2]=='RUN':
		runduty+=1
	if hdadata[i+1].split(',')[2] in ['WD', 'PRT', 'NIGHT']:
		fixedsignon+=1
	for j in range(11):
		ndasheet.cell(row=13+i,column=j+1).value=ndadata[i+1].split(',')[j]
	if ndadata[i+1].split(',')[6]=='Y':
		A+=1
	if ndadata[i+1].split(',')[7]=='Y':
		B+=1
	if ndadata[i+1].split(',')[8]=='Y':
		C+=1
	if ndadata[i+1].split(',')[9]=='Y':
		D+=1
hdasheet['F44']=totalkm; hdasheet['F45']=runduty; hdasheet['F46']=fixedsignon; hdasheet['F47']=runduty+fixedsignon
ndasheet['G44']=A; ndasheet['H44']=B; ndasheet['I44']=C; ndasheet['J44']=D;
hdanda.save(month+' HDA NDA.xlsx')
print('Excel created ')